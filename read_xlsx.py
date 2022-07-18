from openpyxl import load_workbook


def test_some_cell_value(file):
    wb = load_workbook(file)

    sheet = wb.active
    assert sheet.cell(row=3, column=2).value == 'Mara', 'Имя в ячейке не Mara'